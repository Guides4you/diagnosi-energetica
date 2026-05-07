"""
Web App per Diagnosi Energetiche
Streamlit application per la gestione completa delle diagnosi energetiche.

Avvio:
    streamlit run app_diagnosi.py
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import os
import json
import io

# Path delle pagine fisse (template del rapporto, sempre allegate al PDF)
PAGINE_FISSE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pagine_fisse.pdf')

# Configurazione pagina
st.set_page_config(
    page_title="Diagnosi Energetica",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Costanti — fattori di conversione TEP (D.Lgs. 102/2014, FIRE)
TEP_EE = 0.000187
TEP_GAS = 0.000836
TEP_GASOLIO = 0.00086

# Fattori emissioni CO2
CO2_EE = 0.460  # kg/kWh
CO2_GAS = 1.950  # kg/Smc
CO2_GASOLIO = 2.650  # kg/litro

# Poteri calorifici inferiori (per conversione kWh termici → unità vettore)
PCI_GAS_KWH_PER_SMC = 9.97   # kWh/Smc (gas naturale)
PCI_GASOLIO_KWH_PER_L = 10.0  # kWh/litro

MESI = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno",
        "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]

# Modalità test: ?test=true nell'URL carica dati fittizi
MODO_TEST = st.query_params.get("test", "").lower() == "true"

# Inizializzazione session state
if 'anagrafica' not in st.session_state:
    if MODO_TEST:
        st.session_state.anagrafica = {
            'ragione_sociale': 'Industria Esempio S.r.l.',
            'piva': '01234567890',
            'indirizzo': 'Via Roma 42',
            'citta': 'Milano',
            'cap': '20100',
            'provincia': 'MI',
            'ateco': '25.11.00',
            'anno_rif': datetime.now().year - 1,
            'giorni_lav': 250,
            'turni': 1,
            'ore_turno': 8,
        }
    else:
        st.session_state.anagrafica = {
            'ragione_sociale': '',
            'piva': '',
            'indirizzo': '',
            'citta': '',
            'cap': '',
            'provincia': '',
            'ateco': '',
            'anno_rif': datetime.now().year - 1,
            'giorni_lav': 250,
            'turni': 1,
            'ore_turno': 8,
        }

if 'consumi_ee' not in st.session_state:
    if MODO_TEST:
        st.session_state.consumi_ee = pd.DataFrame({
            'Mese': MESI,
            'kWh': [12500.0, 11800.0, 13200.0, 14500.0, 16000.0, 18500.0,
                    19200.0, 17800.0, 15600.0, 14200.0, 13000.0, 12700.0],
            'Costo (€)': [4375.0, 4130.0, 4620.0, 5075.0, 5600.0, 6475.0,
                          6720.0, 6230.0, 5460.0, 4970.0, 4550.0, 4445.0],
        })
    else:
        st.session_state.consumi_ee = pd.DataFrame({
            'Mese': MESI,
            'kWh': [0.0] * 12,
            'Costo (€)': [0.0] * 12,
        })

if 'consumi_gas' not in st.session_state:
    if MODO_TEST:
        st.session_state.consumi_gas = pd.DataFrame({
            'Mese': MESI,
            'Smc': [3500.0, 3200.0, 2800.0, 1500.0, 800.0, 400.0,
                    300.0, 300.0, 600.0, 1200.0, 2500.0, 3400.0],
            'Costo (€)': [3500.0, 3200.0, 2800.0, 1500.0, 800.0, 400.0,
                          300.0, 300.0, 600.0, 1200.0, 2500.0, 3400.0],
        })
    else:
        st.session_state.consumi_gas = pd.DataFrame({
            'Mese': MESI,
            'Smc': [0.0] * 12,
            'Costo (€)': [0.0] * 12,
        })

if 'consumi_gasolio' not in st.session_state:
    if MODO_TEST:
        st.session_state.consumi_gasolio = pd.DataFrame({
            'Mese': MESI,
            'Litri': [500.0, 450.0, 400.0, 200.0, 100.0, 50.0,
                      50.0, 50.0, 100.0, 250.0, 400.0, 500.0],
            'Costo (€)': [750.0, 675.0, 600.0, 300.0, 150.0, 75.0,
                          75.0, 75.0, 150.0, 375.0, 600.0, 750.0],
        })
    else:
        st.session_state.consumi_gasolio = pd.DataFrame({
            'Mese': MESI,
            'Litri': [0.0] * 12,
            'Costo (€)': [0.0] * 12,
        })

if 'bilancio' not in st.session_state:
    if MODO_TEST:
        st.session_state.bilancio = pd.DataFrame({
            'Categoria': ['ATTIVITA PRINCIPALI'] * 3 + ['SERVIZI AUSILIARI'] * 2 + ['SERVIZI GENERALI'] * 2,
            'Descrizione': ['Linea produzione 1', 'Linea produzione 2', 'Magazzino', 'Compressore', 'Pompe', 'Illuminazione', 'Climatizzazione'],
            'Vettore': ['Energia Elettrica', 'Energia Elettrica', 'Gas Naturale', 'Energia Elettrica', 'Energia Elettrica', 'Energia Elettrica', 'Gas Naturale'],
            'Potenza (kW)': [45.0, 30.0, 15.0, 22.0, 7.5, 12.0, 35.0],
            'Ore/giorno': [8.0, 8.0, 8.0, 8.0, 6.0, 10.0, 8.0],
            'Giorni/anno': [250, 250, 250, 250, 250, 300, 200],
            'Fattore carico': [0.70, 0.65, 0.40, 0.60, 0.50, 0.80, 0.55],
            'C.C.': [1.0, 1.0, 0.8, 0.9, 0.7, 1.0, 1.0],
        })
    else:
        st.session_state.bilancio = pd.DataFrame({
            'Categoria': ['ATTIVITA PRINCIPALI'] * 3 + ['SERVIZI AUSILIARI'] * 2 + ['SERVIZI GENERALI'] * 2,
            'Descrizione': ['Reparto 1', 'Reparto 2', 'Reparto 3', 'Compressore', 'Altro', 'Illuminazione', 'Clima'],
            'Vettore': ['Energia Elettrica'] * 7,
            'Potenza (kW)': [0.0] * 7,
            'Ore/giorno': [8.0] * 7,
            'Giorni/anno': [250] * 7,
            'Fattore carico': [0.5] * 7,
            'C.C.': [1.0] * 7,
        })

if 'interventi' not in st.session_state:
    if MODO_TEST:
        st.session_state.interventi = [
            {
                'nome': 'Sostituzione illuminazione con LED',
                'vettore': 'Energia Elettrica',
                'costo_inv': 15000.0,
                'costo_man': 200.0,
                'risparmio': 18000.0,
                'risparmio_euro': 6100.0,
                'vita_utile': 15,
                'tasso': 0.04,
                'payback': 2.5,
                'van': 52800.0,
                'note': 'Sostituzione corpi illuminanti con LED ad alta efficienza',
            },
            {
                'nome': 'Installazione inverter compressore',
                'vettore': 'Energia Elettrica',
                'costo_inv': 8000.0,
                'costo_man': 100.0,
                'risparmio': 8800.0,
                'risparmio_euro': 2980.0,
                'vita_utile': 10,
                'tasso': 0.04,
                'payback': 2.7,
                'van': 16160.0,
                'note': 'Inverter su compressore aria compressa da 22 kW',
            },
        ]
    else:
        st.session_state.interventi = []

if 'driver_energetici' not in st.session_state:
    if MODO_TEST:
        st.session_state.driver_energetici = pd.DataFrame({
            'Nome': ['Superficie totale', 'Superficie riscaldata', 'Produzione', 'Addetti'],
            'Quantità': [1500.0, 800.0, 200.0, 25.0],
            'Unità': ['mq', 'mq', 'ton', 'n'],
            'Categoria': ['Generale', 'Generale', 'Principale', 'Generale'],
        })
    else:
        st.session_state.driver_energetici = pd.DataFrame({
            'Nome': ['Superficie totale'],
            'Quantità': [0.0],
            'Unità': ['mq'],
            'Categoria': ['Generale'],
        })

if 'indici_consumi' not in st.session_state:
    if MODO_TEST:
        st.session_state.indici_consumi = pd.DataFrame({
            'Attività': ['Riscaldamento ambienti', 'Linea produzione 1', 'Illuminazione'],
            'Categoria': ['SERVIZI GENERALI', 'ATTIVITA PRINCIPALI', 'SERVIZI GENERALI'],
            'Vettore': ['Gas Naturale', 'Energia Elettrica', 'Energia Elettrica'],
            'Consumo annuo': [3000.0, 80000.0, 15000.0],
            'Unità consumo': ['Smc', 'kWh', 'kWh'],
            'Driver': ['Superficie riscaldata', 'Produzione', 'Superficie totale'],
        })
    else:
        st.session_state.indici_consumi = pd.DataFrame({
            'Attività': pd.Series([], dtype='object'),
            'Categoria': pd.Series([], dtype='object'),
            'Vettore': pd.Series([], dtype='object'),
            'Consumo annuo': pd.Series([], dtype='float'),
            'Unità consumo': pd.Series([], dtype='object'),
            'Driver': pd.Series([], dtype='object'),
        })

if 'fotovoltaico' not in st.session_state:
    if MODO_TEST:
        st.session_state.fotovoltaico = {
            'potenza_kwp': 50.0,
            'anno_installazione': 2022,
            'produzione_annua': 57500.0,
            'autoconsumo_perc': 70,
        }
    else:
        st.session_state.fotovoltaico = {
            'potenza_kwp': 0.0,
            'anno_installazione': 0,
            'produzione_annua': 0.0,
            'autoconsumo_perc': 70,
        }


def calcola_totali():
    """Calcola i totali energetici."""
    ee_kwh = st.session_state.consumi_ee['kWh'].sum()
    ee_costo = st.session_state.consumi_ee['Costo (€)'].sum()
    gas_smc = st.session_state.consumi_gas['Smc'].sum()
    gas_costo = st.session_state.consumi_gas['Costo (€)'].sum()
    gasolio_l = st.session_state.consumi_gasolio['Litri'].sum()
    gasolio_costo = st.session_state.consumi_gasolio['Costo (€)'].sum()

    return {
        'ee_kwh': ee_kwh,
        'ee_tep': ee_kwh * TEP_EE,
        'ee_costo': ee_costo,
        'ee_co2': ee_kwh * CO2_EE / 1000,
        'gas_smc': gas_smc,
        'gas_tep': gas_smc * TEP_GAS,
        'gas_costo': gas_costo,
        'gas_co2': gas_smc * CO2_GAS / 1000,
        'gasolio_l': gasolio_l,
        'gasolio_tep': gasolio_l * TEP_GASOLIO,
        'gasolio_costo': gasolio_costo,
        'gasolio_co2': gasolio_l * CO2_GASOLIO / 1000,
        'tep_totale': ee_kwh * TEP_EE + gas_smc * TEP_GAS + gasolio_l * TEP_GASOLIO,
        'costo_totale': ee_costo + gas_costo + gasolio_costo,
        'co2_totale': ee_kwh * CO2_EE / 1000 + gas_smc * CO2_GAS / 1000 + gasolio_l * CO2_GASOLIO / 1000,
    }


def calcola_bilancio():
    """Calcola i consumi dal bilancio energetico per ogni vettore."""
    df = st.session_state.bilancio.copy()
    if 'C.C.' not in df.columns:
        df['C.C.'] = 1.0
    if 'Vettore' not in df.columns:
        df['Vettore'] = 'Energia Elettrica'

    df['kWh/anno'] = df['Potenza (kW)'] * df['Ore/giorno'] * df['Giorni/anno'] * df['Fattore carico'] * df['C.C.']

    def _conv(row):
        v = row['Vettore']
        kwh = row['kWh/anno']
        if v == 'Gas Naturale':
            consumo = kwh / PCI_GAS_KWH_PER_SMC
            return pd.Series({'Consumo': consumo, 'Unità': 'Smc', 'TEP': consumo * TEP_GAS, 'CO2 (ton)': consumo * CO2_GAS / 1000})
        elif v == 'Gasolio':
            consumo = kwh / PCI_GASOLIO_KWH_PER_L
            return pd.Series({'Consumo': consumo, 'Unità': 'litri', 'TEP': consumo * TEP_GASOLIO, 'CO2 (ton)': consumo * CO2_GASOLIO / 1000})
        else:  # Energia Elettrica o Altro
            return pd.Series({'Consumo': kwh, 'Unità': 'kWh', 'TEP': kwh * TEP_EE, 'CO2 (ton)': kwh * CO2_EE / 1000})

    df[['Consumo', 'Unità', 'TEP', 'CO2 (ton)']] = df.apply(_conv, axis=1)
    return df


# Sidebar - Navigazione
st.sidebar.title("⚡ Diagnosi Energetica")

# Anno di riferimento sempre visibile
st.sidebar.selectbox(
    "📅 Anno di riferimento",
    options=list(range(2015, datetime.now().year + 1)),
    index=list(range(2015, datetime.now().year + 1)).index(st.session_state.anagrafica['anno_rif']),
    key="anno_sidebar",
    on_change=lambda: st.session_state.anagrafica.update({'anno_rif': st.session_state.anno_sidebar})
)

st.sidebar.markdown("---")

menu = st.sidebar.radio(
    "Sezioni",
    ["🏠 Home", "🏢 Anagrafica", "📊 Consumi EE", "🔥 Consumi Gas",
     "⛽ Consumi Gasolio", "☀️ Fotovoltaico", "⚖️ Bilancio Energetico",
     "📐 Indici Energetici", "🔧 Interventi", "📈 Dashboard", "📄 Genera Report"]
)

st.sidebar.markdown("---")
st.sidebar.markdown("### Riepilogo rapido")
totali = calcola_totali()
st.sidebar.metric("Energia totale", f"{totali['tep_totale']:.2f} TEP")
st.sidebar.metric("Spesa totale", f"€ {totali['costo_totale']:,.0f}")
st.sidebar.metric("CO2 totale", f"{totali['co2_totale']:.1f} ton")


# === SALVA / CARICA PROGETTO (sempre visibile) ===
st.sidebar.markdown("---")
st.sidebar.markdown("### 💾 Progetto")

if st.sidebar.button("💾 Salva progetto", use_container_width=True):
    progetto = {
        'anagrafica': st.session_state.anagrafica,
        'consumi_ee': st.session_state.consumi_ee.to_dict(),
        'consumi_gas': st.session_state.consumi_gas.to_dict(),
        'consumi_gasolio': st.session_state.consumi_gasolio.to_dict(),
        'bilancio': st.session_state.bilancio.to_dict(),
        'driver_energetici': st.session_state.driver_energetici.to_dict(),
        'indici_consumi': st.session_state.indici_consumi.to_dict(),
        'interventi': st.session_state.interventi,
        'fotovoltaico': st.session_state.fotovoltaico,
    }
    _nome = (st.session_state.anagrafica.get('ragione_sociale') or 'progetto').replace(' ', '_').replace('.', '')
    st.session_state.json_data = json.dumps(progetto, indent=2).encode('utf-8')
    st.session_state.json_filename = f"progetto_{_nome}.json"

if 'json_data' in st.session_state:
    st.sidebar.download_button(
        label="⬇️ Scarica file progetto",
        data=st.session_state.json_data,
        file_name=st.session_state.json_filename,
        mime="application/json",
        use_container_width=True,
    )

uploaded = st.sidebar.file_uploader("Carica progetto (.json)", type="json", key="sidebar_uploader")
if uploaded is not None:
    progetto = json.load(uploaded)
    st.session_state.anagrafica = progetto['anagrafica']
    st.session_state.consumi_ee = pd.DataFrame(progetto['consumi_ee'])
    st.session_state.consumi_gas = pd.DataFrame(progetto['consumi_gas'])
    st.session_state.consumi_gasolio = pd.DataFrame(progetto['consumi_gasolio'])
    st.session_state.bilancio = pd.DataFrame(progetto['bilancio'])
    st.session_state.interventi = progetto['interventi']
    st.session_state.fotovoltaico = progetto['fotovoltaico']
    # Compatibilità retroattiva: i campi nuovi possono mancare nei JSON vecchi
    if 'driver_energetici' in progetto:
        st.session_state.driver_energetici = pd.DataFrame(progetto['driver_energetici'])
    if 'indici_consumi' in progetto:
        st.session_state.indici_consumi = pd.DataFrame(progetto['indici_consumi'])
    # Reset chiavi widget così verranno reinizializzate dai nuovi dati
    for k in list(st.session_state.keys()):
        if isinstance(k, str) and (k.startswith('anag_') or k.startswith('fv_') or k.startswith('int_') or k == 'anno_sidebar' or k.startswith('indici_')):
            del st.session_state[k]
    st.sidebar.success("✓ Progetto caricato!")
    st.rerun()


# === PAGINE ===

if menu == "🏠 Home":
    st.title("⚡ Diagnosi Energetica")
    st.markdown("### Sistema di gestione diagnosi energetiche ai sensi del D.Lgs. 102/2014")

    st.markdown("""
    Benvenuto nel sistema di gestione delle diagnosi energetiche.
    Questa applicazione ti permette di:

    - 📝 **Inserire i dati anagrafici** dell'azienda
    - 📊 **Registrare i consumi** di energia elettrica, gas e gasolio
    - ☀️ **Gestire impianti fotovoltaici** e autoproduzione
    - ⚖️ **Creare il bilancio energetico** per area/reparto
    - 🔧 **Proporre interventi** di efficientamento con analisi economica
    - 📈 **Visualizzare dashboard** interattive
    - 📄 **Generare report** Excel e PDF professionali
    """)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.info(f"**Anno di riferimento**\n\n{st.session_state.anagrafica['anno_rif']}")
    with col2:
        azienda = st.session_state.anagrafica['ragione_sociale'] or "Non definita"
        st.info(f"**Azienda**\n\n{azienda}")
    with col3:
        st.info(f"**Consumi totali**\n\n{totali['tep_totale']:.2f} TEP")

    st.markdown("---")
    st.markdown("### Come iniziare")
    st.markdown("""
    1. Vai in **Anagrafica** e inserisci i dati dell'azienda
    2. Inserisci i **Consumi** mensili (EE, Gas, Gasolio)
    3. Compila il **Bilancio Energetico** per area
    4. Aggiungi gli **Interventi** di efficientamento
    5. Controlla la **Dashboard** per una visione d'insieme
    6. **Genera il Report** finale
    """)


elif menu == "🏢 Anagrafica":
    st.title("🏢 Dati Anagrafici")

    # Init chiavi widget dai valori del dict (solo al primo accesso)
    for k in ['ragione_sociale', 'piva', 'indirizzo', 'citta', 'cap', 'provincia', 'ateco',
              'anno_rif', 'giorni_lav', 'turni', 'ore_turno']:
        wkey = f"anag_{k}"
        if wkey not in st.session_state:
            st.session_state[wkey] = st.session_state.anagrafica.get(k, '' if k in ['ragione_sociale','piva','indirizzo','citta','cap','provincia','ateco'] else 0)

    col1, col2 = st.columns(2)

    with col1:
        st.text_input("Ragione Sociale *", key="anag_ragione_sociale")
        st.text_input("Partita IVA *", key="anag_piva")
        st.text_input("Indirizzo", key="anag_indirizzo")
        st.text_input("Città", key="anag_citta")

    with col2:
        st.text_input("CAP", key="anag_cap")
        st.text_input("Provincia", key="anag_provincia")
        st.text_input("Codice ATECO", key="anag_ateco")
        st.number_input(
            "Anno di riferimento *",
            min_value=2015,
            max_value=datetime.now().year,
            key="anag_anno_rif",
        )

    st.markdown("### Regime operativo")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.number_input("Giorni lavorativi/anno", min_value=1, max_value=365, key="anag_giorni_lav")
    with col2:
        st.number_input("Numero turni", min_value=1, max_value=3, key="anag_turni")
    with col3:
        st.number_input("Ore per turno", min_value=1, max_value=12, key="anag_ore_turno")

    # Sync widget → dict
    for k in ['ragione_sociale', 'piva', 'indirizzo', 'citta', 'cap', 'provincia', 'ateco',
              'anno_rif', 'giorni_lav', 'turni', 'ore_turno']:
        st.session_state.anagrafica[k] = st.session_state[f"anag_{k}"]

    if st.session_state.anagrafica['ragione_sociale']:
        st.success(f"✓ Anagrafica salvata per: {st.session_state.anagrafica['ragione_sociale']}")


elif menu == "📊 Consumi EE":
    st.title("📊 Consumi Energia Elettrica")

    st.markdown("Inserisci i consumi mensili di energia elettrica.")

    # Editor tabella
    edited_df = st.data_editor(
        st.session_state.consumi_ee,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "Mese": st.column_config.TextColumn("Mese", disabled=True),
            "kWh": st.column_config.NumberColumn("kWh", min_value=0, format="%.0f"),
            "Costo (€)": st.column_config.NumberColumn("Costo (€)", min_value=0, format="%.2f"),
        }
    )
    st.session_state.consumi_ee = edited_df

    # Totali
    tot_kwh = edited_df['kWh'].sum()
    tot_costo = edited_df['Costo (€)'].sum()
    costo_medio = tot_costo / tot_kwh if tot_kwh > 0 else 0

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Totale kWh", f"{tot_kwh:,.0f}")
    col2.metric("Totale €", f"{tot_costo:,.2f}")
    col3.metric("TEP", f"{tot_kwh * TEP_EE:.2f}")
    col4.metric("€/kWh medio", f"{costo_medio:.4f}")

    # Grafico
    if tot_kwh > 0:
        fig = px.bar(edited_df, x='Mese', y='kWh', title='Consumi mensili EE')
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)


elif menu == "🔥 Consumi Gas":
    st.title("🔥 Consumi Gas Naturale")

    st.markdown("Inserisci i consumi mensili di gas naturale.")

    edited_df = st.data_editor(
        st.session_state.consumi_gas,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "Mese": st.column_config.TextColumn("Mese", disabled=True),
            "Smc": st.column_config.NumberColumn("Smc", min_value=0, format="%.0f"),
            "Costo (€)": st.column_config.NumberColumn("Costo (€)", min_value=0, format="%.2f"),
        }
    )
    st.session_state.consumi_gas = edited_df

    tot_smc = edited_df['Smc'].sum()
    tot_costo = edited_df['Costo (€)'].sum()

    col1, col2, col3 = st.columns(3)
    col1.metric("Totale Smc", f"{tot_smc:,.0f}")
    col2.metric("Totale €", f"{tot_costo:,.2f}")
    col3.metric("TEP", f"{tot_smc * TEP_GAS:.2f}")

    if tot_smc > 0:
        fig = px.bar(edited_df, x='Mese', y='Smc', title='Consumi mensili Gas')
        fig.update_layout(xaxis_tickangle=-45)
        st.plotly_chart(fig, use_container_width=True)


elif menu == "⛽ Consumi Gasolio":
    st.title("⛽ Consumi Gasolio")

    st.markdown("Inserisci i consumi mensili di gasolio.")

    edited_df = st.data_editor(
        st.session_state.consumi_gasolio,
        num_rows="fixed",
        use_container_width=True,
        column_config={
            "Mese": st.column_config.TextColumn("Mese", disabled=True),
            "Litri": st.column_config.NumberColumn("Litri", min_value=0, format="%.0f"),
            "Costo (€)": st.column_config.NumberColumn("Costo (€)", min_value=0, format="%.2f"),
        }
    )
    st.session_state.consumi_gasolio = edited_df

    tot_litri = edited_df['Litri'].sum()
    tot_costo = edited_df['Costo (€)'].sum()

    col1, col2, col3 = st.columns(3)
    col1.metric("Totale Litri", f"{tot_litri:,.0f}")
    col2.metric("Totale €", f"{tot_costo:,.2f}")
    col3.metric("TEP", f"{tot_litri * TEP_GASOLIO:.2f}")


elif menu == "☀️ Fotovoltaico":
    st.title("☀️ Impianto Fotovoltaico")

    st.markdown("Inserisci i dati dell'impianto fotovoltaico (se presente).")

    # Init chiavi widget dai valori del dict
    if 'fv_potenza_kwp' not in st.session_state:
        st.session_state.fv_potenza_kwp = float(st.session_state.fotovoltaico.get('potenza_kwp', 0.0))
    if 'fv_anno_installazione' not in st.session_state:
        anno = st.session_state.fotovoltaico.get('anno_installazione', 0)
        st.session_state.fv_anno_installazione = int(anno) if anno else datetime.now().year
    if 'fv_produzione_annua' not in st.session_state:
        st.session_state.fv_produzione_annua = float(st.session_state.fotovoltaico.get('produzione_annua', 0.0))
    if 'fv_autoconsumo_perc' not in st.session_state:
        st.session_state.fv_autoconsumo_perc = int(st.session_state.fotovoltaico.get('autoconsumo_perc', 70))

    col1, col2 = st.columns(2)
    with col1:
        st.number_input(
            "Potenza installata (kWp)",
            min_value=0.0, max_value=10000.0, step=0.1,
            key="fv_potenza_kwp",
        )
        st.number_input(
            "Anno installazione",
            min_value=2000, max_value=datetime.now().year,
            key="fv_anno_installazione",
        )

    with col2:
        st.number_input(
            "Produzione annua (kWh)",
            min_value=0.0, step=100.0,
            key="fv_produzione_annua",
        )
        st.slider(
            "Percentuale autoconsumo (%)",
            min_value=0, max_value=100,
            key="fv_autoconsumo_perc",
        )

    # Sync widget → dict
    st.session_state.fotovoltaico['potenza_kwp'] = st.session_state.fv_potenza_kwp
    st.session_state.fotovoltaico['anno_installazione'] = st.session_state.fv_anno_installazione
    st.session_state.fotovoltaico['produzione_annua'] = st.session_state.fv_produzione_annua
    st.session_state.fotovoltaico['autoconsumo_perc'] = st.session_state.fv_autoconsumo_perc

    # Stima produzione se non inserita
    if st.session_state.fotovoltaico['potenza_kwp'] > 0 and st.session_state.fotovoltaico['produzione_annua'] == 0:
        stima = st.session_state.fotovoltaico['potenza_kwp'] * 1150  # Nord Italia
        st.info(f"💡 Produzione stimata (Nord Italia): {stima:,.0f} kWh/anno")
        if st.button("Usa stima"):
            st.session_state.fv_produzione_annua = stima
            st.session_state.fotovoltaico['produzione_annua'] = stima
            st.rerun()

    if st.session_state.fotovoltaico['produzione_annua'] > 0:
        autoconsumo = st.session_state.fotovoltaico['produzione_annua'] * st.session_state.fotovoltaico['autoconsumo_perc'] / 100
        immissione = st.session_state.fotovoltaico['produzione_annua'] - autoconsumo

        col1, col2, col3 = st.columns(3)
        col1.metric("Produzione", f"{st.session_state.fotovoltaico['produzione_annua']:,.0f} kWh")
        col2.metric("Autoconsumo", f"{autoconsumo:,.0f} kWh")
        col3.metric("Immissione rete", f"{immissione:,.0f} kWh")


elif menu == "⚖️ Bilancio Energetico":
    st.title("⚖️ Bilancio Energetico")

    st.markdown("Definisci la ripartizione dei consumi per area/reparto e per vettore energetico.")

    # Migrazione automatica per progetti vecchi
    if 'C.C.' not in st.session_state.bilancio.columns:
        st.session_state.bilancio['C.C.'] = 1.0
    if 'Vettore' not in st.session_state.bilancio.columns:
        st.session_state.bilancio['Vettore'] = 'Energia Elettrica'

    # Riordino colonne per UX (Vettore subito dopo Descrizione)
    col_order = ['Categoria', 'Descrizione', 'Vettore', 'Potenza (kW)', 'Ore/giorno', 'Giorni/anno', 'Fattore carico', 'C.C.']
    st.session_state.bilancio = st.session_state.bilancio[[c for c in col_order if c in st.session_state.bilancio.columns]]

    # Editor
    edited_df = st.data_editor(
        st.session_state.bilancio,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Categoria": st.column_config.SelectboxColumn(
                "Categoria",
                options=["ATTIVITA PRINCIPALI", "SERVIZI AUSILIARI", "SERVIZI GENERALI", "ALTRO"]
            ),
            "Descrizione": st.column_config.TextColumn("Descrizione"),
            "Vettore": st.column_config.SelectboxColumn(
                "Vettore",
                options=["Energia Elettrica", "Gas Naturale", "Gasolio", "Altro"],
                help="Vettore energetico associato all'utenza",
            ),
            "Potenza (kW)": st.column_config.NumberColumn("Potenza (kW)", min_value=0, format="%.2f", help="Potenza elettrica nominale (EE) o termica al focolare (Gas/Gasolio)"),
            "Ore/giorno": st.column_config.NumberColumn("Ore/giorno", min_value=0, max_value=24, format="%.1f"),
            "Giorni/anno": st.column_config.NumberColumn("Giorni/anno", min_value=0, max_value=365),
            "Fattore carico": st.column_config.NumberColumn("Fc (carico)", min_value=0, max_value=1, format="%.2f", help="Fattore di carico: frazione di potenza media rispetto alla nominale"),
            "C.C.": st.column_config.NumberColumn("C.C. (contemp.)", min_value=0, max_value=1, format="%.2f", help="Fattore di contemporaneità: frazione di tempo in cui l'utenza è attiva contemporaneamente"),
        }
    )
    st.session_state.bilancio = edited_df

    st.caption(f"ℹ️ Conversione: kWh termici → Smc usando PCI gas {PCI_GAS_KWH_PER_SMC} kWh/Smc; → litri usando PCI gasolio {PCI_GASOLIO_KWH_PER_L} kWh/litro.")

    # Calcoli
    df_calc = calcola_bilancio()

    st.markdown("### Risultati")

    totale_kwh = df_calc['kWh/anno'].sum()
    totale_tep = df_calc['TEP'].sum()

    # Riepilogo doppio: per Categoria e per Vettore
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Per categoria**")
        riepilogo_cat = df_calc.groupby('Categoria').agg({'kWh/anno': 'sum', 'TEP': 'sum'}).reset_index()
        st.dataframe(riepilogo_cat, use_container_width=True, hide_index=True)
        if totale_kwh > 0:
            fig = px.pie(riepilogo_cat, values='kWh/anno', names='Categoria', title='Ripartizione per categoria (kWh)')
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown("**Per vettore**")
        riepilogo_vet = df_calc.groupby(['Vettore', 'Unità']).agg({'Consumo': 'sum', 'TEP': 'sum'}).reset_index()
        st.dataframe(riepilogo_vet, use_container_width=True, hide_index=True)
        if totale_tep > 0:
            fig = px.pie(riepilogo_vet, values='TEP', names='Vettore', title='Ripartizione per vettore (TEP)')
            st.plotly_chart(fig, use_container_width=True)

    col_a, col_b = st.columns(2)
    col_a.metric("Energia totale (kWh equivalenti)", f"{totale_kwh:,.0f}")
    col_b.metric("TEP totale", f"{totale_tep:.2f}")

    # Confronto con consumi reali per vettore
    st.markdown("### Confronto con consumi da bollette")

    bollette = {
        'Energia Elettrica': (st.session_state.consumi_ee['kWh'].sum(), 'kWh'),
        'Gas Naturale': (st.session_state.consumi_gas['Smc'].sum(), 'Smc'),
        'Gasolio': (st.session_state.consumi_gasolio['Litri'].sum(), 'litri'),
    }

    for vettore, (reale, unita) in bollette.items():
        df_v = df_calc[df_calc['Vettore'] == vettore]
        stimato = df_v['Consumo'].sum() if not df_v.empty else 0.0
        if stimato == 0 and reale == 0:
            continue
        c1, c2, c3, c4 = st.columns([1.2, 1, 1, 1])
        c1.write(f"**{vettore}**")
        c2.metric("Bilancio", f"{stimato:,.0f} {unita}")
        c3.metric("Bollette", f"{reale:,.0f} {unita}")
        if reale > 0:
            diff_perc = (stimato - reale) / reale * 100
            c4.metric("Δ", f"{diff_perc:+.1f}%")
            if abs(diff_perc) > 20:
                st.warning(f"⚠️ {vettore}: scostamento bilancio/bollette > 20%. Verifica i dati o considera autoproduzione (es. fotovoltaico).")


elif menu == "📐 Indici Energetici":
    st.title("📐 Indici Energetici")
    st.markdown("Definisci i **driver** (mq, kg, ton, pz...) e i **consumi per attività**. L'app calcola gli indici di prestazione (es. kWh/ton, Smc/mq).")

    # --- DRIVER
    st.markdown("### 1️⃣ Driver di consumo")
    st.caption("I driver sono le grandezze che 'spiegano' il consumo: superficie, produzione, addetti, ore di funzionamento, ecc.")

    driver_df = st.data_editor(
        st.session_state.driver_energetici,
        num_rows="dynamic",
        use_container_width=True,
        key="indici_driver_editor",
        column_config={
            "Nome": st.column_config.TextColumn("Nome", help="Es: 'Superficie riscaldata', 'Produzione pane'"),
            "Quantità": st.column_config.NumberColumn("Quantità", min_value=0.0, format="%.2f"),
            "Unità": st.column_config.SelectboxColumn(
                "Unità",
                options=["mq", "mc", "kg", "ton", "pz", "n", "ore", "addetti", "litri"],
            ),
            "Categoria": st.column_config.SelectboxColumn(
                "Categoria",
                options=["Principale", "Ausiliario", "Generale"],
                help="Tipo di attività che il driver caratterizza",
            ),
        },
    )
    st.session_state.driver_energetici = driver_df

    # --- CONSUMI
    st.markdown("### 2️⃣ Consumi per attività")
    st.caption("Associa il consumo annuo di ogni attività al driver di riferimento.")

    driver_options = sorted([n for n in driver_df['Nome'].dropna().tolist() if str(n).strip()])
    if not driver_options:
        driver_options = ['—']

    indici_df = st.data_editor(
        st.session_state.indici_consumi,
        num_rows="dynamic",
        use_container_width=True,
        key="indici_consumi_editor",
        column_config={
            "Attività": st.column_config.TextColumn("Attività", help="Es: 'Cottura', 'Climatizzazione uffici'"),
            "Categoria": st.column_config.SelectboxColumn(
                "Categoria",
                options=["ATTIVITA PRINCIPALI", "SERVIZI AUSILIARI", "SERVIZI GENERALI"],
            ),
            "Vettore": st.column_config.SelectboxColumn(
                "Vettore",
                options=["Energia Elettrica", "Gas Naturale", "Gasolio", "Altro"],
            ),
            "Consumo annuo": st.column_config.NumberColumn("Consumo annuo", min_value=0.0, format="%.2f"),
            "Unità consumo": st.column_config.SelectboxColumn(
                "Unità consumo",
                options=["kWh", "Smc", "litri", "MJ", "TEP"],
            ),
            "Driver": st.column_config.SelectboxColumn(
                "Driver",
                options=driver_options,
                help="Driver associato all'attività",
            ),
        },
    )
    st.session_state.indici_consumi = indici_df

    # --- CALCOLO INDICI
    st.markdown("### 3️⃣ Indici calcolati")

    rows = []
    if not indici_df.empty and not driver_df.empty:
        for _, r in indici_df.iterrows():
            dname = str(r.get('Driver') or '').strip()
            if not dname or dname == '—':
                continue
            d_match = driver_df[driver_df['Nome'] == dname]
            if d_match.empty:
                continue
            quantita = float(d_match.iloc[0]['Quantità'] or 0)
            if quantita == 0:
                continue
            unita_d = d_match.iloc[0]['Unità']
            consumo = float(r['Consumo annuo'] or 0)
            unita_c = r['Unità consumo']
            indice = consumo / quantita
            rows.append({
                'Categoria': r['Categoria'],
                'Attività': r['Attività'],
                'Vettore': r['Vettore'],
                'Indice': indice,
                'Unità indice': f"{unita_c}/{unita_d}",
                'Calcolo': f"{consumo:,.1f} {unita_c} ÷ {quantita:,.1f} {unita_d}",
            })

    if rows:
        df_idx = pd.DataFrame(rows)
        st.session_state.indici_calcolati = df_idx

        for cat in ["ATTIVITA PRINCIPALI", "SERVIZI AUSILIARI", "SERVIZI GENERALI"]:
            group = df_idx[df_idx['Categoria'] == cat]
            if group.empty:
                continue
            st.markdown(f"**{cat}**")
            st.dataframe(
                group[['Attività', 'Vettore', 'Indice', 'Unità indice', 'Calcolo']],
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Indice': st.column_config.NumberColumn('Valore', format="%.4f"),
                },
            )

        # Riepilogo numero indici
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        col1.metric("Indici totali", len(df_idx))
        col2.metric("Driver utilizzati", len(set(indici_df['Driver'].dropna()) & set(driver_df['Nome'].dropna())))
        col3.metric("Categorie coperte", df_idx['Categoria'].nunique())
    else:
        st.session_state.indici_calcolati = pd.DataFrame()
        st.info("ℹ️ Per calcolare gli indici servono almeno un driver con Quantità > 0 e un'attività con Driver associato.")


elif menu == "🔧 Interventi":
    st.title("🔧 Interventi di Efficientamento")

    st.markdown("Aggiungi le proposte di intervento di efficientamento energetico.")

    # Form nuovo intervento
    with st.expander("➕ Aggiungi nuovo intervento", expanded=len(st.session_state.interventi) == 0):
        col1, col2 = st.columns(2)

        with col1:
            nome = st.text_input("Nome intervento", key="int_nome")
            vettore = st.selectbox("Vettore risparmiato", ["Energia Elettrica", "Gas Naturale", "Gasolio"], key="int_vettore")
            costo_inv = st.number_input("Costo investimento (€)", min_value=0.0, step=100.0, key="int_costo_inv")
            costo_man = st.number_input("Costo manutenzione annuo (€)", min_value=0.0, step=10.0, key="int_costo_man")

        with col2:
            risparmio = st.number_input("Risparmio energetico annuo (kWh/Smc/l)", min_value=0.0, step=100.0, key="int_risparmio")
            vita_utile = st.number_input("Vita utile (anni)", min_value=1, max_value=30, value=10, key="int_vita_utile")
            tasso = st.number_input("Tasso attualizzazione (%)", min_value=0.0, max_value=20.0, value=4.0, key="int_tasso") / 100

            # Costo medio vettore
            if vettore == "Energia Elettrica":
                tot_kwh = st.session_state.consumi_ee['kWh'].sum()
                tot_costo = st.session_state.consumi_ee['Costo (€)'].sum()
                costo_medio = tot_costo / tot_kwh if tot_kwh > 0 else 0.36
            elif vettore == "Gas Naturale":
                tot = st.session_state.consumi_gas['Smc'].sum()
                tot_costo = st.session_state.consumi_gas['Costo (€)'].sum()
                costo_medio = tot_costo / tot if tot > 0 else 1.0
            else:
                tot = st.session_state.consumi_gasolio['Litri'].sum()
                tot_costo = st.session_state.consumi_gasolio['Costo (€)'].sum()
                costo_medio = tot_costo / tot if tot > 0 else 1.5

            st.info(f"Costo medio {vettore}: € {costo_medio:.4f}")

        note = st.text_area("Note", key="int_note")

        if st.button("Aggiungi intervento"):
            if nome and costo_inv > 0 and risparmio > 0:
                # Calcolo payback e VAN
                risparmio_euro = risparmio * costo_medio - costo_man
                payback = costo_inv / risparmio_euro if risparmio_euro > 0 else 999

                # VAN
                van = -costo_inv
                for anno in range(1, vita_utile + 1):
                    van += risparmio_euro / ((1 + tasso) ** anno)

                intervento = {
                    'nome': nome,
                    'vettore': vettore,
                    'costo_inv': costo_inv,
                    'costo_man': costo_man,
                    'risparmio': risparmio,
                    'risparmio_euro': risparmio_euro,
                    'vita_utile': vita_utile,
                    'tasso': tasso,
                    'payback': payback,
                    'van': van,
                    'note': note,
                }
                st.session_state.interventi.append(intervento)
                st.success(f"✓ Intervento '{nome}' aggiunto!")
                # Reset form
                for k in ['int_nome', 'int_costo_inv', 'int_costo_man', 'int_risparmio', 'int_note']:
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()
            else:
                st.error("Compila tutti i campi obbligatori (nome, costo, risparmio)")

    # Lista interventi
    if st.session_state.interventi:
        st.markdown("### Interventi proposti")

        for i, interv in enumerate(st.session_state.interventi):
            with st.container():
                col1, col2, col3, col4, col5 = st.columns([3, 1, 1, 1, 0.5])
                col1.write(f"**{interv['nome']}**")
                col2.write(f"€ {interv['costo_inv']:,.0f}")
                col3.write(f"€ {interv['risparmio_euro']:,.0f}/anno")
                col4.write(f"PB: {interv['payback']:.1f}a")

                van_color = "green" if interv['van'] > 0 else "red"
                col4.markdown(f"VAN: <span style='color:{van_color}'>€ {interv['van']:,.0f}</span>", unsafe_allow_html=True)

                if col5.button("🗑️", key=f"del_{i}"):
                    st.session_state.interventi.pop(i)
                    st.rerun()

        # Riepilogo
        st.markdown("---")
        tot_inv = sum(i['costo_inv'] for i in st.session_state.interventi)
        tot_risp = sum(i['risparmio_euro'] for i in st.session_state.interventi)
        tot_van = sum(i['van'] for i in st.session_state.interventi)

        col1, col2, col3 = st.columns(3)
        col1.metric("Investimento totale", f"€ {tot_inv:,.0f}")
        col2.metric("Risparmio totale", f"€ {tot_risp:,.0f}/anno")
        col3.metric("VAN totale", f"€ {tot_van:,.0f}")


elif menu == "📈 Dashboard":
    st.title("📈 Dashboard Energetica")

    totali = calcola_totali()

    # KPI principali
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Energia Totale", f"{totali['tep_totale']:.2f} TEP")
    col2.metric("Spesa Totale", f"€ {totali['costo_totale']:,.0f}")
    col3.metric("Emissioni CO2", f"{totali['co2_totale']:.1f} ton")
    if totali['ee_kwh'] > 0:
        col4.metric("Costo medio EE", f"€ {totali['ee_costo']/totali['ee_kwh']:.4f}/kWh")

    st.markdown("---")

    # Grafici
    col1, col2 = st.columns(2)

    with col1:
        # Ripartizione per vettore (TEP)
        data_tep = pd.DataFrame({
            'Vettore': ['Energia Elettrica', 'Gas Naturale', 'Gasolio'],
            'TEP': [totali['ee_tep'], totali['gas_tep'], totali['gasolio_tep']]
        })
        data_tep = data_tep[data_tep['TEP'] > 0]

        if not data_tep.empty:
            fig = px.pie(data_tep, values='TEP', names='Vettore', title='Ripartizione consumi (TEP)')
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        # Ripartizione costi
        data_costo = pd.DataFrame({
            'Vettore': ['Energia Elettrica', 'Gas Naturale', 'Gasolio'],
            'Costo': [totali['ee_costo'], totali['gas_costo'], totali['gasolio_costo']]
        })
        data_costo = data_costo[data_costo['Costo'] > 0]

        if not data_costo.empty:
            fig = px.pie(data_costo, values='Costo', names='Vettore', title='Ripartizione spesa (€)')
            st.plotly_chart(fig, use_container_width=True)

    # Andamento mensile
    st.markdown("### Andamento mensile consumi")

    df_mensile = pd.DataFrame({
        'Mese': MESI,
        'Energia Elettrica (kWh)': st.session_state.consumi_ee['kWh'],
        'Gas Naturale (Smc)': st.session_state.consumi_gas['Smc'],
    })

    fig = go.Figure()
    fig.add_trace(go.Bar(name='EE (kWh)', x=df_mensile['Mese'], y=df_mensile['Energia Elettrica (kWh)']))
    fig.add_trace(go.Bar(name='Gas (Smc)', x=df_mensile['Mese'], y=df_mensile['Gas Naturale (Smc)']))
    fig.update_layout(barmode='group', xaxis_tickangle=-45)
    st.plotly_chart(fig, use_container_width=True)

    # Bilancio energetico
    df_bil = calcola_bilancio()
    if df_bil['kWh/anno'].sum() > 0:
        st.markdown("### Bilancio energetico per categoria")
        riepilogo = df_bil.groupby('Categoria')['kWh/anno'].sum().reset_index()
        fig = px.bar(riepilogo, x='Categoria', y='kWh/anno', title='Consumi per categoria')
        st.plotly_chart(fig, use_container_width=True)


elif menu == "📄 Genera Report":
    st.title("📄 Genera Report")

    st.markdown("Genera i file di output della diagnosi energetica.")

    # Verifica dati minimi
    totali = calcola_totali()
    has_anagrafica = bool(st.session_state.anagrafica['ragione_sociale'])
    has_consumi = totali['tep_totale'] > 0

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### Checklist")
        st.checkbox("Anagrafica compilata", value=has_anagrafica, disabled=True)
        st.checkbox("Consumi inseriti", value=has_consumi, disabled=True)
        st.checkbox("Bilancio energetico", value=calcola_bilancio()['kWh/anno'].sum() > 0, disabled=True)
        st.checkbox("Interventi proposti", value=len(st.session_state.interventi) > 0, disabled=True)

    with col2:
        st.markdown("### Riepilogo")
        st.write(f"**Azienda:** {st.session_state.anagrafica['ragione_sociale'] or 'N/D'}")
        st.write(f"**Anno:** {st.session_state.anagrafica['anno_rif']}")
        st.write(f"**Energia totale:** {totali['tep_totale']:.2f} TEP")
        st.write(f"**Spesa totale:** € {totali['costo_totale']:,.0f}")
        st.write(f"**Interventi:** {len(st.session_state.interventi)}")

    st.markdown("---")

    if has_anagrafica and has_consumi:
        nome_azienda = st.session_state.anagrafica['ragione_sociale'].replace(' ', '_').replace('.', '')

        col1, col2 = st.columns(2)

        # --- GENERA EXCEL ---
        with col1:
            if st.button("📊 Genera Excel", use_container_width=True):
                from openpyxl import Workbook
                from openpyxl.styles import Font

                wb = Workbook()

                # Foglio SINTESI
                ws = wb.active
                ws.title = "SINTESI"
                ws['A1'] = f"DIAGNOSI ENERGETICA - {st.session_state.anagrafica['ragione_sociale']}"
                ws['A1'].font = Font(bold=True, size=14)
                ws['A2'] = f"Anno di riferimento: {st.session_state.anagrafica['anno_rif']}"

                ws['A4'] = "Vettore"
                ws['B4'] = "Consumo"
                ws['C4'] = "Unità"
                ws['D4'] = "TEP"
                ws['E4'] = "Costo (€)"

                ws['A5'] = "Energia Elettrica"
                ws['B5'] = totali['ee_kwh']
                ws['C5'] = "kWh"
                ws['D5'] = totali['ee_tep']
                ws['E5'] = totali['ee_costo']

                ws['A6'] = "Gas Naturale"
                ws['B6'] = totali['gas_smc']
                ws['C6'] = "Smc"
                ws['D6'] = totali['gas_tep']
                ws['E6'] = totali['gas_costo']

                ws['A7'] = "Gasolio"
                ws['B7'] = totali['gasolio_l']
                ws['C7'] = "litri"
                ws['D7'] = totali['gasolio_tep']
                ws['E7'] = totali['gasolio_costo']

                ws['A8'] = "TOTALE"
                ws['D8'] = totali['tep_totale']
                ws['E8'] = totali['costo_totale']

                # Foglio CONSUMI_EE
                ws_ee = wb.create_sheet("CONSUMI_EE")
                ws_ee['A1'] = "Mese"
                ws_ee['B1'] = "kWh"
                ws_ee['C1'] = "Costo (€)"
                for i, row in st.session_state.consumi_ee.iterrows():
                    ws_ee[f'A{i+2}'] = row['Mese']
                    ws_ee[f'B{i+2}'] = row['kWh']
                    ws_ee[f'C{i+2}'] = row['Costo (€)']

                # Foglio BILANCIO (esteso multi-vettore)
                ws_bil = wb.create_sheet("BILANCIO")
                df_bil = calcola_bilancio()
                bil_cols = ["Categoria", "Descrizione", "Vettore", "Potenza (kW)", "Ore/giorno", "Giorni/anno", "Fattore carico", "C.C.", "kWh/anno", "Consumo", "Unità", "TEP", "CO2 (ton)"]
                for j, col in enumerate(bil_cols):
                    ws_bil.cell(row=1, column=j+1, value=col).font = Font(bold=True)
                for i, row in df_bil.iterrows():
                    for j, col in enumerate(bil_cols):
                        ws_bil.cell(row=i+2, column=j+1, value=row.get(col))

                # Foglio DRIVER + INDICI
                if not st.session_state.driver_energetici.empty:
                    ws_drv = wb.create_sheet("DRIVER")
                    drv_cols = list(st.session_state.driver_energetici.columns)
                    for j, col in enumerate(drv_cols):
                        ws_drv.cell(row=1, column=j+1, value=col).font = Font(bold=True)
                    for i, row in st.session_state.driver_energetici.iterrows():
                        for j, col in enumerate(drv_cols):
                            ws_drv.cell(row=i+2, column=j+1, value=row.get(col))

                if 'indici_calcolati' in st.session_state and not st.session_state.indici_calcolati.empty:
                    ws_idx = wb.create_sheet("INDICI")
                    idx_df = st.session_state.indici_calcolati
                    idx_cols = ["Categoria", "Attività", "Vettore", "Indice", "Unità indice", "Calcolo"]
                    for j, col in enumerate(idx_cols):
                        ws_idx.cell(row=1, column=j+1, value=col).font = Font(bold=True)
                    for i, row in idx_df.iterrows():
                        for j, col in enumerate(idx_cols):
                            ws_idx.cell(row=i+2, column=j+1, value=row.get(col))

                # Foglio INTERVENTI
                if st.session_state.interventi:
                    ws_int = wb.create_sheet("INTERVENTI")
                    int_cols = ["Intervento", "Vettore", "Investimento (€)", "Manutenzione/anno (€)", "Risparmio energia/anno", "Risparmio €/anno", "Vita utile (anni)", "Tasso (%)", "Payback (anni)", "VAN (€)", "Note"]
                    for j, col in enumerate(int_cols):
                        ws_int.cell(row=1, column=j+1, value=col).font = Font(bold=True)
                    for i, interv in enumerate(st.session_state.interventi):
                        ws_int.cell(row=i+2, column=1, value=interv['nome'])
                        ws_int.cell(row=i+2, column=2, value=interv['vettore'])
                        ws_int.cell(row=i+2, column=3, value=interv['costo_inv'])
                        ws_int.cell(row=i+2, column=4, value=interv['costo_man'])
                        ws_int.cell(row=i+2, column=5, value=interv['risparmio'])
                        ws_int.cell(row=i+2, column=6, value=interv['risparmio_euro'])
                        ws_int.cell(row=i+2, column=7, value=interv['vita_utile'])
                        ws_int.cell(row=i+2, column=8, value=interv['tasso'] * 100)
                        ws_int.cell(row=i+2, column=9, value=interv['payback'])
                        ws_int.cell(row=i+2, column=10, value=interv['van'])
                        ws_int.cell(row=i+2, column=11, value=interv.get('note', ''))

                # Salva in memoria (BytesIO)
                buffer_excel = io.BytesIO()
                wb.save(buffer_excel)
                buffer_excel.seek(0)

                st.session_state.excel_data = buffer_excel.getvalue()
                st.session_state.excel_filename = f"DIAGNOSI_{nome_azienda}_{st.session_state.anagrafica['anno_rif']}.xlsx"
                st.success("✓ File Excel generato!")

        # --- GENERA PDF ---
        with col2:
            if st.button("📄 Genera PDF", use_container_width=True):
                try:
                    from reportlab.lib import colors
                    from reportlab.lib.pagesizes import A4
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import cm
                    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
                    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
                    from pypdf import PdfReader, PdfWriter

                    # Tracker pagine sezioni dinamiche
                    section_pages_dyn = {}

                    class TrackingDoc(SimpleDocTemplate):
                        def afterFlowable(self, flowable):
                            if hasattr(flowable, '_bookmark') and flowable._bookmark not in section_pages_dyn:
                                section_pages_dyn[flowable._bookmark] = self.page

                    def tag(para, name):
                        para._bookmark = name
                        return para

                    styles = getSampleStyleSheet()
                    styles.add(ParagraphStyle(name='TitoloReport', parent=styles['Title'],
                                            fontSize=20, textColor=colors.HexColor('#1a5276'),
                                            alignment=TA_CENTER))
                    styles.add(ParagraphStyle(name='Body', parent=styles['Normal'],
                                            fontSize=10, alignment=TA_JUSTIFY, spaceAfter=6))
                    styles.add(ParagraphStyle(name='SezTitolo', parent=styles['Heading1'],
                                            fontSize=14, textColor=colors.HexColor('#1a5276'),
                                            spaceAfter=12, spaceBefore=6))
                    styles.add(ParagraphStyle(name='IndiceRiga', parent=styles['Normal'],
                                            fontSize=11, alignment=TA_LEFT, spaceAfter=4, leftIndent=10))

                    PRIMARY = colors.HexColor('#2874a6')
                    LIGHT = colors.HexColor('#d5e8d4')

                    def make_table(data, col_widths, total_row=False, font_size=9):
                        t = Table(data, colWidths=col_widths, repeatRows=1)
                        style = [
                            ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                            ('FONTSIZE', (0, 0), (-1, -1), font_size),
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f7')]),
                        ]
                        if total_row:
                            style.append(('BACKGROUND', (0, -1), (-1, -1), LIGHT))
                            style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
                        t.setStyle(TableStyle(style))
                        return t

                    anag = st.session_state.anagrafica
                    story_body = []

                    # ===== SEZIONE 9 — DATI ANAGRAFICI =====
                    story_body.append(tag(Paragraph("9. DATI ANAGRAFICI E REGIME OPERATIVO", styles['SezTitolo']), 'anagrafica'))
                    data_an = [
                        ['Campo', 'Valore'],
                        ['Ragione Sociale', anag.get('ragione_sociale', '—')],
                        ['Partita IVA', anag.get('piva', '—')],
                        ['Sede', f"{anag.get('indirizzo','')}, {anag.get('cap','')} {anag.get('citta','')} ({anag.get('provincia','')})"],
                        ['Codice ATECO', anag.get('ateco', '—')],
                        ['Anno riferimento', str(anag.get('anno_rif', ''))],
                        ['Giorni lavorativi/anno', str(anag.get('giorni_lav', ''))],
                        ['Numero turni', str(anag.get('turni', ''))],
                        ['Ore per turno', str(anag.get('ore_turno', ''))],
                    ]
                    story_body.append(make_table(data_an, [5*cm, 11*cm]))
                    story_body.append(Spacer(1, 0.8*cm))

                    # ===== SEZIONE 10 — VETTORI ENERGETICI =====
                    story_body.append(PageBreak())
                    story_body.append(tag(Paragraph("10. QUADRO DEI VETTORI ENERGETICI", styles['SezTitolo']), 'vettori'))
                    data_sint = [
                        ['Vettore', 'Consumo', 'TEP', 'CO₂ (ton)', 'Costo (€)'],
                        ['Energia Elettrica', f"{totali['ee_kwh']:,.0f} kWh", f"{totali['ee_tep']:.2f}", f"{totali['ee_co2']:.2f}", f"{totali['ee_costo']:,.0f}"],
                        ['Gas Naturale', f"{totali['gas_smc']:,.0f} Smc", f"{totali['gas_tep']:.2f}", f"{totali['gas_co2']:.2f}", f"{totali['gas_costo']:,.0f}"],
                        ['Gasolio', f"{totali['gasolio_l']:,.0f} litri", f"{totali['gasolio_tep']:.2f}", f"{totali['gasolio_co2']:.2f}", f"{totali['gasolio_costo']:,.0f}"],
                        ['TOTALE', '', f"{totali['tep_totale']:.2f}", f"{totali['co2_totale']:.2f}", f"{totali['costo_totale']:,.0f}"],
                    ]
                    story_body.append(make_table(data_sint, [4.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 3*cm], total_row=True))
                    story_body.append(Spacer(1, 0.6*cm))

                    # Tabelle mensili dettagliate
                    def _tab_mensile(df_consumi, label_quant, totale_q, totale_c):
                        rows = [['Mese', label_quant, 'Costo (€)']]
                        for _, r in df_consumi.iterrows():
                            rows.append([r['Mese'], f"{r[label_quant]:,.0f}", f"{r['Costo (€)']:,.2f}"])
                        rows.append(['TOTALE', f"{totale_q:,.0f}", f"{totale_c:,.2f}"])
                        return make_table(rows, [4*cm, 4*cm, 4*cm], total_row=True, font_size=8)

                    if totali['ee_kwh'] > 0:
                        story_body.append(Paragraph("Dettaglio mensile — Energia Elettrica", styles['Heading3']))
                        story_body.append(_tab_mensile(st.session_state.consumi_ee, 'kWh', totali['ee_kwh'], totali['ee_costo']))
                        story_body.append(Spacer(1, 0.4*cm))
                    if totali['gas_smc'] > 0:
                        story_body.append(Paragraph("Dettaglio mensile — Gas Naturale", styles['Heading3']))
                        story_body.append(_tab_mensile(st.session_state.consumi_gas, 'Smc', totali['gas_smc'], totali['gas_costo']))
                        story_body.append(Spacer(1, 0.4*cm))
                    if totali['gasolio_l'] > 0:
                        story_body.append(Paragraph("Dettaglio mensile — Gasolio", styles['Heading3']))
                        story_body.append(_tab_mensile(st.session_state.consumi_gasolio, 'Litri', totali['gasolio_l'], totali['gasolio_costo']))
                        story_body.append(Spacer(1, 0.4*cm))

                    # ===== SEZIONE 11 — FOTOVOLTAICO =====
                    fv = st.session_state.fotovoltaico
                    has_fv = fv.get('potenza_kwp', 0) > 0
                    if has_fv:
                        story_body.append(PageBreak())
                        story_body.append(tag(Paragraph("11. IMPIANTO FOTOVOLTAICO", styles['SezTitolo']), 'fotovoltaico'))
                        autoconsumo_kwh = fv['produzione_annua'] * fv['autoconsumo_perc'] / 100
                        immissione_kwh = fv['produzione_annua'] - autoconsumo_kwh
                        data_fv = [
                            ['Parametro', 'Valore'],
                            ['Potenza installata', f"{fv['potenza_kwp']:.2f} kWp"],
                            ['Anno installazione', str(fv.get('anno_installazione', '—'))],
                            ['Produzione annua', f"{fv['produzione_annua']:,.0f} kWh"],
                            ['Autoconsumo', f"{fv['autoconsumo_perc']}% — {autoconsumo_kwh:,.0f} kWh"],
                            ['Immissione in rete', f"{immissione_kwh:,.0f} kWh"],
                        ]
                        story_body.append(make_table(data_fv, [6*cm, 10*cm]))
                        story_body.append(Spacer(1, 0.6*cm))

                    # ===== SEZIONE 12 — BILANCIO ENERGETICO =====
                    df_bil = calcola_bilancio()
                    has_bil = (not df_bil.empty) and df_bil['kWh/anno'].sum() > 0
                    if has_bil:
                        story_body.append(PageBreak())
                        story_body.append(tag(Paragraph("12. BILANCIO ENERGETICO", styles['SezTitolo']), 'bilancio'))

                        # Dettaglio per area
                        story_body.append(Paragraph("Dettaglio per utenza", styles['Heading3']))
                        data_b = [['Categoria', 'Descrizione', 'Vettore', 'kWh/anno', 'Consumo', 'TEP']]
                        for _, r in df_bil.iterrows():
                            data_b.append([
                                str(r.get('Categoria', '')),
                                str(r.get('Descrizione', ''))[:25],
                                str(r.get('Vettore', ''))[:15],
                                f"{r['kWh/anno']:,.0f}",
                                f"{r['Consumo']:,.1f} {r['Unità']}",
                                f"{r['TEP']:.2f}",
                            ])
                        story_body.append(make_table(data_b, [3.5*cm, 3.5*cm, 3*cm, 2.2*cm, 2.5*cm, 1.8*cm], font_size=8))
                        story_body.append(Spacer(1, 0.5*cm))

                        # Per categoria
                        story_body.append(Paragraph("Riepilogo per categoria", styles['Heading3']))
                        riep_cat = df_bil.groupby('Categoria').agg({'kWh/anno': 'sum', 'TEP': 'sum'}).reset_index()
                        rows_c = [['Categoria', 'kWh/anno', 'TEP']]
                        for _, r in riep_cat.iterrows():
                            rows_c.append([r['Categoria'], f"{r['kWh/anno']:,.0f}", f"{r['TEP']:.2f}"])
                        rows_c.append(['TOTALE', f"{df_bil['kWh/anno'].sum():,.0f}", f"{df_bil['TEP'].sum():.2f}"])
                        story_body.append(make_table(rows_c, [7*cm, 4*cm, 3*cm], total_row=True))
                        story_body.append(Spacer(1, 0.4*cm))

                        # Per vettore
                        story_body.append(Paragraph("Riepilogo per vettore", styles['Heading3']))
                        riep_vet = df_bil.groupby(['Vettore', 'Unità']).agg({'Consumo': 'sum', 'TEP': 'sum'}).reset_index()
                        rows_v = [['Vettore', 'Consumo', 'Unità', 'TEP']]
                        for _, r in riep_vet.iterrows():
                            rows_v.append([r['Vettore'], f"{r['Consumo']:,.1f}", r['Unità'], f"{r['TEP']:.2f}"])
                        story_body.append(make_table(rows_v, [5*cm, 4*cm, 2.5*cm, 3*cm]))
                        story_body.append(Spacer(1, 0.5*cm))

                    # ===== SEZIONE 13 — INDICI ENERGETICI =====
                    has_idx = 'indici_calcolati' in st.session_state and not st.session_state.indici_calcolati.empty
                    if has_idx:
                        story_body.append(PageBreak())
                        story_body.append(tag(Paragraph("13. INDICI DI PRESTAZIONE ENERGETICA", styles['SezTitolo']), 'indici'))
                        df_idx = st.session_state.indici_calcolati
                        for cat in ["ATTIVITA PRINCIPALI", "SERVIZI AUSILIARI", "SERVIZI GENERALI"]:
                            group = df_idx[df_idx['Categoria'] == cat]
                            if group.empty:
                                continue
                            story_body.append(Paragraph(cat, styles['Heading3']))
                            rows_i = [['Attività', 'Vettore', 'Valore', 'Unità', 'Calcolo']]
                            for _, r in group.iterrows():
                                rows_i.append([
                                    str(r['Attività'])[:25],
                                    str(r['Vettore'])[:15],
                                    f"{r['Indice']:.4f}",
                                    str(r['Unità indice']),
                                    str(r['Calcolo'])[:30],
                                ])
                            story_body.append(make_table(rows_i, [4*cm, 3*cm, 2.5*cm, 2.5*cm, 4*cm], font_size=8))
                            story_body.append(Spacer(1, 0.4*cm))

                    # ===== SEZIONE 14 — INTERVENTI =====
                    has_int = bool(st.session_state.interventi)
                    if has_int:
                        story_body.append(PageBreak())
                        story_body.append(tag(Paragraph("14. INTERVENTI DI EFFICIENTAMENTO ENERGETICO", styles['SezTitolo']), 'interventi'))

                        # Tabella sintetica
                        rows_int = [['Intervento', 'Vettore', 'Investim.', 'Risparmio €/a', 'PB (a)', 'VAN (€)']]
                        for interv in st.session_state.interventi:
                            rows_int.append([
                                str(interv['nome'])[:30],
                                str(interv['vettore'])[:12],
                                f"€ {interv['costo_inv']:,.0f}",
                                f"€ {interv['risparmio_euro']:,.0f}",
                                f"{interv['payback']:.1f}",
                                f"€ {interv['van']:,.0f}",
                            ])
                        story_body.append(make_table(rows_int, [4.5*cm, 2.5*cm, 2.3*cm, 2.5*cm, 1.5*cm, 2.5*cm], font_size=8))
                        story_body.append(Spacer(1, 0.5*cm))

                        # Dettaglio per ogni intervento
                        for idx, interv in enumerate(st.session_state.interventi, 1):
                            story_body.append(Paragraph(f"Intervento {idx}: {interv['nome']}", styles['Heading3']))
                            data_d = [
                                ['Voce', 'Valore'],
                                ['Vettore risparmiato', interv['vettore']],
                                ['Costo investimento', f"€ {interv['costo_inv']:,.2f}"],
                                ['Costo manutenzione annuo', f"€ {interv['costo_man']:,.2f}"],
                                ['Risparmio energetico annuo', f"{interv['risparmio']:,.0f} (kWh/Smc/litri)"],
                                ['Risparmio economico annuo', f"€ {interv['risparmio_euro']:,.2f}"],
                                ['Vita utile', f"{interv['vita_utile']} anni"],
                                ['Tasso attualizzazione', f"{interv['tasso']*100:.1f}%"],
                                ['Tempo di ritorno (Payback)', f"{interv['payback']:.1f} anni"],
                                ['VAN', f"€ {interv['van']:,.2f}"],
                            ]
                            story_body.append(make_table(data_d, [6*cm, 10*cm], font_size=9))
                            if interv.get('note'):
                                story_body.append(Spacer(1, 0.2*cm))
                                story_body.append(Paragraph(f"<i>Note:</i> {interv['note']}", styles['Body']))
                            story_body.append(Spacer(1, 0.4*cm))

                    # ===== SEZIONE 15 — CONCLUSIONI =====
                    story_body.append(PageBreak())
                    story_body.append(tag(Paragraph("15. CONCLUSIONI E SINTESI", styles['SezTitolo']), 'conclusioni'))

                    story_body.append(Paragraph(
                        f"L'azienda <b>{anag.get('ragione_sociale','')}</b> nell'anno di riferimento <b>{anag.get('anno_rif','')}</b> ha consumato complessivamente <b>{totali['tep_totale']:.2f} TEP</b> "
                        f"per una spesa totale di <b>€ {totali['costo_totale']:,.0f}</b>, con emissioni stimate di "
                        f"<b>{totali['co2_totale']:.2f} tonnellate</b> di CO₂.",
                        styles['Body']
                    ))

                    if st.session_state.interventi:
                        tot_inv = sum(i['costo_inv'] for i in st.session_state.interventi)
                        tot_risp = sum(i['risparmio_euro'] for i in st.session_state.interventi)
                        tot_van = sum(i['van'] for i in st.session_state.interventi)
                        pb_glob = tot_inv / tot_risp if tot_risp > 0 else 0
                        story_body.append(Spacer(1, 0.3*cm))
                        story_body.append(Paragraph(
                            f"Sono stati proposti <b>{len(st.session_state.interventi)} interventi</b> di efficientamento, "
                            f"per un investimento complessivo di <b>€ {tot_inv:,.0f}</b> e un risparmio annuo di "
                            f"<b>€ {tot_risp:,.0f}</b> (Payback medio {pb_glob:.1f} anni, VAN totale € {tot_van:,.0f}).",
                            styles['Body']
                        ))

                    story_body.append(Spacer(1, 0.6*cm))
                    story_body.append(Paragraph("Riepilogo finale", styles['Heading3']))
                    rows_fin = [
                        ['Indicatore', 'Valore'],
                        ['Energia totale', f"{totali['tep_totale']:.2f} TEP"],
                        ['Spesa energetica totale', f"€ {totali['costo_totale']:,.2f}"],
                        ['Emissioni CO₂', f"{totali['co2_totale']:.2f} ton/anno"],
                    ]
                    if st.session_state.interventi:
                        rows_fin.append(['Investimenti proposti', f"€ {sum(i['costo_inv'] for i in st.session_state.interventi):,.0f}"])
                        rows_fin.append(['Risparmio annuo atteso', f"€ {sum(i['risparmio_euro'] for i in st.session_state.interventi):,.0f}"])
                    story_body.append(make_table(rows_fin, [8*cm, 8*cm]))

                    story_body.append(Spacer(1, 1*cm))
                    story_body.append(Paragraph(
                        "<i>Il presente report è stato redatto ai sensi del D.Lgs. 102/2014 e della norma UNI CEI EN 16247. "
                        "I dati riportati sono ricavati dalle bollette dell'anno di riferimento e dalle stime di bilancio energetico.</i>",
                        styles['Body']
                    ))

                    # ===== STEP 1: build del corpo dinamico (per tracciare numeri pagina) =====
                    buffer_body = io.BytesIO()
                    doc_body = TrackingDoc(buffer_body, pagesize=A4,
                                          rightMargin=2*cm, leftMargin=2*cm,
                                          topMargin=2*cm, bottomMargin=2*cm)
                    doc_body.build(story_body)

                    # Numero di pagine fisse e calcolo offset (Cover=1, Indice=1, Pagine fisse=N_fisse)
                    n_fisse = 0
                    if os.path.exists(PAGINE_FISSE_PATH):
                        try:
                            n_fisse = len(PdfReader(PAGINE_FISSE_PATH).pages)
                        except Exception:
                            n_fisse = 0

                    OFFSET_DYN = 2 + n_fisse  # cover + indice + pagine fisse

                    # ===== STEP 2: build copertina + indice =====
                    buffer_cover = io.BytesIO()
                    doc_cover = SimpleDocTemplate(buffer_cover, pagesize=A4,
                                                  rightMargin=2*cm, leftMargin=2*cm,
                                                  topMargin=2*cm, bottomMargin=2*cm)
                    story_cover = []

                    # Copertina
                    story_cover.append(Spacer(1, 3*cm))
                    story_cover.append(Paragraph("DIAGNOSI ENERGETICA", styles['TitoloReport']))
                    story_cover.append(Paragraph("ai sensi del D.Lgs. 102/2014 — UNI CEI EN 16247", styles['Heading2']))
                    story_cover.append(Spacer(1, 2*cm))
                    story_cover.append(Paragraph(f"<b>{anag.get('ragione_sociale','')}</b>", styles['TitoloReport']))
                    story_cover.append(Paragraph(f"{anag.get('indirizzo','')}", styles['Normal']))
                    story_cover.append(Paragraph(f"{anag.get('cap','')} {anag.get('citta','')} ({anag.get('provincia','')})", styles['Normal']))
                    if anag.get('piva'):
                        story_cover.append(Paragraph(f"P.IVA: {anag.get('piva','')}", styles['Normal']))
                    if anag.get('ateco'):
                        story_cover.append(Paragraph(f"Codice ATECO: {anag.get('ateco','')}", styles['Normal']))
                    story_cover.append(Spacer(1, 2*cm))
                    story_cover.append(Paragraph(f"Anno di riferimento: <b>{anag.get('anno_rif','')}</b>", styles['Heading2']))
                    story_cover.append(Spacer(1, 4*cm))
                    story_cover.append(Paragraph(f"Documento generato il {datetime.now().strftime('%d/%m/%Y')}", styles['Normal']))
                    story_cover.append(PageBreak())

                    # Indice
                    story_cover.append(Paragraph("INDICE", styles['SezTitolo']))

                    # Voci pagine fisse (note dal contenuto del PDF template)
                    # I numeri di pagina partono da 3 (dopo cover + indice)
                    indice_rows = [
                        ['#', 'Sezione', 'Pagina'],
                        ['', 'Copertina', '1'],
                        ['', 'Indice', '2'],
                    ]
                    if n_fisse > 0:
                        # Mappa sezioni delle pagine fisse — pagina = 3 + (offset interno)
                        fisse_voci = [
                            ('1.', 'Premessa', 3),
                            ('2.', 'Overview', 4),
                            ('3.', 'Glossario', 9),
                            ('4.', 'Unità di misura', 13),
                            ('5.', 'Gruppo Finservice', 14),
                            ('6.', 'Informazioni', 15),
                            ('7.', 'Fase 2 — Audit', 16),
                            ('8.', 'Fase 3 — Reporting', 17),
                        ]
                        for num, titolo, p in fisse_voci:
                            if p <= 2 + n_fisse:
                                indice_rows.append([num, titolo, str(p)])

                    # Voci sezioni dinamiche (numeri pagina dal tracking + offset)
                    voci_dyn = [
                        ('9.', 'Dati Anagrafici e Regime Operativo', 'anagrafica'),
                        ('10.', 'Quadro dei Vettori Energetici', 'vettori'),
                    ]
                    if has_fv:
                        voci_dyn.append(('11.', 'Impianto Fotovoltaico', 'fotovoltaico'))
                    if has_bil:
                        voci_dyn.append(('12.', 'Bilancio Energetico', 'bilancio'))
                    if has_idx:
                        voci_dyn.append(('13.', 'Indici di Prestazione Energetica', 'indici'))
                    if has_int:
                        voci_dyn.append(('14.', 'Interventi di Efficientamento', 'interventi'))
                    voci_dyn.append(('15.', 'Conclusioni e Sintesi', 'conclusioni'))

                    for num, titolo, key in voci_dyn:
                        p = section_pages_dyn.get(key, 1) + OFFSET_DYN
                        indice_rows.append([num, titolo, str(p)])

                    t_indice = Table(indice_rows, colWidths=[1.5*cm, 12*cm, 2.5*cm], repeatRows=1)
                    t_indice.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f7')]),
                    ]))
                    story_cover.append(t_indice)

                    doc_cover.build(story_cover)

                    # ===== STEP 3: merge finale (cover+indice + pagine fisse + body dinamico) =====
                    writer = PdfWriter()
                    buffer_cover.seek(0)
                    for page in PdfReader(buffer_cover).pages:
                        writer.add_page(page)
                    if n_fisse > 0 and os.path.exists(PAGINE_FISSE_PATH):
                        for page in PdfReader(PAGINE_FISSE_PATH).pages:
                            writer.add_page(page)
                    buffer_body.seek(0)
                    for page in PdfReader(buffer_body).pages:
                        writer.add_page(page)

                    buffer_pdf = io.BytesIO()
                    writer.write(buffer_pdf)
                    buffer_pdf.seek(0)

                    st.session_state.pdf_data = buffer_pdf.getvalue()
                    st.session_state.pdf_filename = f"REPORT_DE_{nome_azienda}_{anag.get('anno_rif','')}.pdf"
                    st.success(f"✓ File PDF generato! ({len(writer.pages)} pagine totali — copertina + indice + {n_fisse} pagine fisse + sezioni dinamiche)")

                except Exception as e:
                    st.error(f"Errore nella generazione del PDF: {e}")

        # Download buttons persistenti (fuori dal st.button)
        st.markdown("---")
        col_dl1, col_dl2 = st.columns(2)

        with col_dl1:
            if 'excel_data' in st.session_state:
                st.download_button(
                    label="⬇️ Scarica Excel",
                    data=st.session_state.excel_data,
                    file_name=st.session_state.excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        with col_dl2:
            if 'pdf_data' in st.session_state:
                st.download_button(
                    label="⬇️ Scarica PDF",
                    data=st.session_state.pdf_data,
                    file_name=st.session_state.pdf_filename,
                    mime="application/pdf",
                    use_container_width=True,
                )
    else:
        st.warning("⚠️ Compila almeno l'anagrafica e inserisci i consumi prima di generare il report.")

    st.info("💾 Per salvare/caricare il progetto usa i pulsanti nella sidebar a sinistra.")


# Footer
st.sidebar.markdown("---")
st.sidebar.markdown("*Diagnosi Energetica v1.0*")
st.sidebar.markdown("*D.Lgs. 102/2014 - UNI CEI EN 16247*")
